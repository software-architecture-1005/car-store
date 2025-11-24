import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from core.interfaces.report_generator import IVehicleReportGenerator

class ExcelReportGenerator(IVehicleReportGenerator):
    """
    Implementación concreta de IVehicleReportGenerator usando openpyxl.
    Esta clase es el 'detalle' de bajo nivel que depende de la abstracción.
    Demuestra el Principio de Inversión de Dependencias (DIP).
    """
    
    def generate_report(self, vehicle) -> bytes:
        # Crear workbook y worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Vehículo"
        
        # Estilos
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=14)
        title_font = Font(bold=True, size=18)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Título
        ws.merge_cells('A1:B1')
        ws['A1'] = f"Reporte de Vehículo: {vehicle.make.name} {vehicle.model}"
        ws['A1'].font = title_font
        ws['A1'].alignment = center_alignment
        ws.row_dimensions[1].height = 30
        
        # Encabezados de tabla
        headers = ["Atributo", "Detalle"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        
        # Datos del vehículo
        data = [
            ["Marca", vehicle.make.name],
            ["Modelo", vehicle.model],
            ["Año", str(vehicle.year)],
            ["Precio", f"${vehicle.price:,.2f}"],
            ["Color", vehicle.color],
            ["Categoría", vehicle.category.name if vehicle.category else "N/A"],
            ["Disponible", "Sí" if vehicle.is_available else "No"]
        ]
        
        # Escribir datos
        for row_num, (attribute, value) in enumerate(data, 4):
            # Columna de atributo
            attr_cell = ws.cell(row=row_num, column=1)
            attr_cell.value = attribute
            attr_cell.font = Font(bold=True)
            attr_cell.border = border
            attr_cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Columna de valor
            value_cell = ws.cell(row=row_num, column=2)
            value_cell.value = value
            value_cell.border = border
            value_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        
        # Footer
        footer_row = len(data) + 5
        ws.merge_cells(f'A{footer_row}:B{footer_row}')
        footer_cell = ws.cell(row=footer_row, column=1)
        footer_cell.value = "Generado automáticamente por Car Store System"
        footer_cell.font = Font(size=8, italic=True, color="808080")
        footer_cell.alignment = center_alignment
        
        # Guardar en buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()


